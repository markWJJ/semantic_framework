# -*- coding: utf-8 -*- 
from api.base import BaseHandler
from openpyxl import load_workbook, Workbook
from settings.setting import database_path, word_path
from io import BytesIO
from guoxin_poc_test import poc_api
import threading
import logging


class ImportHandler(BaseHandler):

    def excel(self, name, body):
        postfix = name.split(".")[-1]
        if postfix == 'xlsx':
            wb = load_workbook(filename=BytesIO(body))
            table = wb.active
            result = [table.cell(row=row, column=1).value for row in range(2, table.max_row + 1)]
            return result
        else:
            return None

    def activate_job(self, data):
        def run_job(self, data):
            logging.info("set block {}".format(self.get_block()))
            length = len(data)
            for idx, line in enumerate(data):
                d={
                    "index_es": False,
                    "store_es": True,
                    "retrive_es": False,
                    "delete_es": False,
                    "es_api": self.es,
                    "database_index": "guoxin_es_database",
                    "semantic_index": "guoxin_semantic_result",
                    # "entity_path": "/data/stock_entities.txt",
                    # "es_database": "/data/es_database.json",
                    "query": line,
                    "data_id": idx,
                }
                try:
                    poc_api(d)
                except:
                    logging.debug("{} -----> 出现异常，已忽略".format(line))
                    pass
                inter = (idx + 1) / length
                self.set_block("%.2f" % inter)
            logging.info("set block {}".format(self.get_block()))

        thread = threading.Thread(target=run_job, args=(self, data,))
        thread.start()
    
    def post(self):
        self.set_block()
        file_meta = self.request.files['file'][0]

        file_name = file_meta.get("filename")
        file_body = file_meta.get("body")
        logging.info('receive file {}'.format(file_name))
        result = self.excel(file_name, file_body)
        if not result:
            self.response({"code": 1, "data": "Unknown file type"})
        else:
            self.activate_job(result)
            self.response(({"code": 0, "data": "success"}))


class IntervalHandler(BaseHandler):

    def get(self):
        block = self.get_block()
        logging.info("check block {}".format(block))

        self.response({"code": 0, "data": block})
